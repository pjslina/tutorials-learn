import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def convert_value(value):
    """尝试把字符串数值转换为 int/float，否则保持原值"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        val = value.strip()
        if re.fullmatch(r"-?\d+", val):  # int
            return int(val)
        if re.fullmatch(r"-?\d+(\.\d+)?", val):  # float
            return float(val)
    return value


def get_display_width(value):
    """估算Excel中字符显示宽度，中文按2个单位宽度，英文和数字按1个"""
    if value is None:
        return 0
    str_val = str(value)
    width = 0
    for char in str_val:
        if re.match(r'[\u4e00-\u9fff]', char):  # 中文
            width += 2
        else:
            width += 1
    return width


def collect_keys_in_order(list_data):
    """按首次出现顺序收集所有键"""
    ordered = []
    seen = set()
    for item in list_data:
        for k in item.keys():
            if k not in seen:
                seen.add(k)
                ordered.append(k)
    return ordered


def ensure_headers_union(ws, list_data, key_column_name=None):
    """
    计算并写回 Sheet 的表头：现有表头 ∪ list_data 的所有键（保持原有顺序，新键追加）
    必要时补齐 key_column_name
    返回最终 headers（list）
    """
    # 读取现有表头（若无则为空）
    existing_headers = []
    if ws.max_row >= 1:
        # 仅取非 None 的头
        max_col = ws.max_column
        for i in range(1, max_col + 1):
            v = ws.cell(row=1, column=i).value
            if v is not None:
                existing_headers.append(v)

    # 收集数据中的所有键（按首次出现顺序）
    data_keys = collect_keys_in_order(list_data)

    # 初始为现有表头
    headers = list(existing_headers)

    # 确保 key 列在表头中
    if key_column_name and key_column_name not in headers:
        headers.append(key_column_name)

    # 追加数据中新出现但表头没有的键
    for k in data_keys:
        if k not in headers:
            headers.append(k)

    # 若 sheet 是空白（A1=None），或者需要扩列/刷新表头，则写回整行表头
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value != header:
            cell.value = header

    return headers


def write_or_update_excel(filepath, sheetname, list_data, key_column_name=None, extra_calc=None, ref_sheet_data=None):
    """
    通用写入或更新 Excel 的函数
    :param filepath: 文件路径
    :param sheetname: sheet 名称
    :param list_data: 要写入的数据（list[dict]）
    :param key_column_name: 主键列名，用于判断是否更新（可为空）
    :param extra_calc: 额外计算逻辑的回调函数，接受 (item, ref_sheet_data) 并返回更新后的 item
    :param ref_sheet_data: 参考 sheet 的原始数据，用于额外计算
    """
    if not list_data:
        # 没数据就直接返回（也可选择创建空表头）
        return

    # 加工数据（如工时计算），确保新增字段也参与表头并集
    if extra_calc:
        list_data = [extra_calc(dict(item), ref_sheet_data) for item in list_data]  # 复制以免修改原引用

    # 打开或创建工作簿/工作表
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb[sheetname] if sheetname in wb.sheetnames else wb.create_sheet(sheetname)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheetname

    # —— 关键修复：表头取并集，并写回 —— #
    headers = ensure_headers_union(ws, list_data, key_column_name=key_column_name)

    # 构建 key -> 行号 的映射（若指定了 key）
    key_to_row = {}
    if key_column_name and key_column_name in headers:
        key_col_idx = headers.index(key_column_name) + 1
        for row_idx in range(2, ws.max_row + 1):
            key = ws.cell(row=row_idx, column=key_col_idx).value
            if key is not None:
                key_to_row[key] = row_idx

    # 写入或更新数据
    for item in list_data:
        # 取 key
        key = item.get(key_column_name) if key_column_name else None

        if key_column_name and key in key_to_row:
            # 更新该行：仅更新 item 中提供的列
            row_idx = key_to_row[key]
            for col_name, value in item.items():
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    new_value = convert_value(value)
                    if ws.cell(row=row_idx, column=col_idx).value != new_value:
                        ws.cell(row=row_idx, column=col_idx, value=new_value)
        else:
            # 新增一行：遍历 headers，存在的填值，不存在的留空
            new_row_idx = ws.max_row + 1
            for col_name in headers:
                col_idx = headers.index(col_name) + 1
                if col_name in item:
                    ws.cell(row=new_row_idx, column=col_idx, value=convert_value(item[col_name]))
                else:
                    # 保持为空（不要写 0 或 ""）
                    pass
            # 若有 key，记录映射
            if key_column_name and key is not None:
                key_to_row[key] = new_row_idx

    # 设置第一行可筛选
    ws.auto_filter.ref = ws.dimensions

    # 冻结前两列和第一行
    ws.freeze_panes = 'C2'

    # 自动设置列宽（最长内容 + 限制最大宽度）
    max_width_limit = 5 * 8.43  # 约42
    min_width = 10
    for col_idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        max_width = get_display_width(header)
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            width = get_display_width(cell_value)
            max_width = max(max_width, width)
        adjusted_width = min(max(max_width * 1.2, min_width), max_width_limit)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 自动换行 + 差值三色标记
    diff_col_idx = headers.index("差值") + 1 if "差值" in headers else None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
        if diff_col_idx:
            diff_cell = row[diff_col_idx - 1]
            if isinstance(diff_cell.value, (int, float)):
                if diff_cell.value < 0:
                    diff_cell.font = Font(color="FF0000")  # 红
                elif diff_cell.value == 0:
                    diff_cell.font = Font(color="00AA00")  # 绿
                else:
                    diff_cell.font = Font(color="0000FF")  # 蓝

    wb.save(filepath)
    print(f"✅ 数据已成功写入或更新：{filepath} 的 sheet《{sheetname}》")


# ========== 示例数据 ==========
listA = [
    {"编号": 1001, "姓名": "张三", "年龄": "28", "成绩": "56.78", "预估人天": "5"},
    {"编号": 1002, "姓名": "李四", "年龄": "30", "成绩": "90", "预估人天": "8"},
    {"编号": 1003, "姓名": "王五", "年龄": "25", "成绩": "67.0", "预估人天": "7"},
    {"编号": 1004, "姓名": "王五", "年龄": "25", "成绩": "55.5", "预估人天": "6"},
    {"编号": 1005, "姓名": "王五", "年龄": "25", "成绩": "78.5", "预估人天": "8"},
]

listB = [
    {"姓名": "张三", "系数": "0.9", "天数": "10", "基线": "1.0"},
    {"姓名": "李四", "系数": "1.0", "天数": "10", "基线": "1.0"},
    {"姓名": "王五", "系数": "1.6", "天数": "10", "基线": "1.0"}
]


# ========== 额外计算逻辑（给工时对比用）==========
def calc_for_sheetB(item, ref_data):
    系数 = convert_value(item.get("系数", 0))
    天数 = convert_value(item.get("天数", 0))
    基线 = convert_value(item.get("基线", 0))

    总天数 = 系数 * 天数 * 基线

    # 计算实际天数（来自“员工数据”的预估人天按姓名汇总）
    实际天数 = 0
    for row in ref_data:
        if row.get("姓名") == item.get("姓名"):
            实际天数 += convert_value(row.get("预估人天", 0))

    差值 = 总天数 - 实际天数

    item["总天数"] = 总天数
    item["实际天数"] = 实际天数
    item["差值"] = 差值
    return item


# ========== 主程序 ==========
file_path = "E:/xml_docs/员工信息.xlsx"

# 先写入/更新 员工数据（会把现有表头与新数据键取并集）
write_or_update_excel(file_path, "员工数据", listA, key_column_name="编号")

# 再写入/更新 工时对比（依赖员工数据；同样表头取并集）
write_or_update_excel(
    file_path,
    "工时对比",
    listB,
    key_column_name="姓名",
    extra_calc=calc_for_sheetB,
    ref_sheet_data=listA
)
